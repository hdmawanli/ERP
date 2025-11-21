import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
import pyautogui
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By


def get_num():
    browser = webdriver.Edge()
    urls = ['http://136.142.46.246:8089/zentao/bug-browse-4.html',
            'http://136.142.46.246:8089/zentao/bug-browse-10.html']
    user = ['fanjihua', 'Hb!@#123456']
    browser.maximize_window()
    browser.implicitly_wait(5)
    browser.get(url=urls[0])
    time.sleep(1)
    browser.find_element(by=By.XPATH,value='//*[@id="account"]').send_keys(user[0])
    time.sleep(0.1)
    browser.find_element(by=By.XPATH,value='//*[@id="loginPanel"]/div/div[2]/form/table/tbody/tr[2]/td/input').send_keys(user[1])
    time.sleep(0.5)
    browser.find_element(by=By.XPATH,value='//*[@id="submit"]').click()
    time.sleep(0.5)
    for url in urls:
        if url == urls[1]:
            browser.get(url=urls[1])
        browser.find_element(by=By.XPATH,value='//*[@id="mainMenu"]/div[3]/div/button').click()
        time.sleep(0.5)
        browser.find_element(by=By.XPATH,value='//*[@id="exportActionMenu"]').click()
        time.sleep(0.5)
        coords = pyautogui.locateOnScreen('img/1.png')
        x, y = pyautogui.center(coords)
        pyautogui.leftClick(x, y)
        time.sleep(2)
    path = os.path.join(os.path.expanduser('~'), 'Downloads')
    Desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    csv_files = [f for f in os.listdir(path) if 'Bug' in f and f.endswith('.csv')]
    for csv_file_name in csv_files:
        shutil.move(os.path.join(path, csv_file_name), Desktop)


def main():
    path = os.path.join(os.path.expanduser('~'), 'Desktop')
    csv_files = [f for f in os.listdir(path) if f.endswith('.csv')]
    for csv_file_name in csv_files:
        print("CSV文件名:", csv_file_name)
        cols_to_use = ['Bug编号', '创建日期', '指派给', 'Bug状态']

        # 读取csv文件
        df = pd.read_csv(os.path.join(path, csv_file_name), usecols=cols_to_use, index_col=False,
                         parse_dates=['创建日期'])
        df['年份'] = df['创建日期'].dt.year
        df['月份'] = df['创建日期'].dt.month
        quarters_dict = {1: '一', 2: '二', 3: '三', 4: '四'}
        df['季度'] = df['创建日期'].dt.quarter.map(lambda q: f'第{quarters_dict[q]}季')
        pivot_table = df[df['Bug状态'] == '处理中'].pivot_table(
            columns=['年份', '季度', '月份'],
            index='指派给',
            values='Bug编号',
            aggfunc='count',
            margins=True,  # 添加行/列总计
            margins_name='总计',
            fill_value=0
        )
        print(pivot_table)
        # 输出透视表到Excel
        excel_file_path = os.path.join(path, csv_file_name.replace('.csv', '.xlsx'))
        with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
            pivot_table.to_excel(writer, sheet_name='Sheet1', index=True)  # 注意这里保留了index  

        # 加载刚刚写入的Excel文件以插入行和合并单元格  
        wb = load_workbook(excel_file_path)
        ws = wb['Sheet1']
        # 在第二行之前插入空行（即第一行之后）  
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ws.max_column)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = Font(bold=False)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(bold=True)
        ws.cell(row=1, column=1).value = csv_file_name.replace('.csv', '')

        # 保存工作簿
        wb.save(excel_file_path)


if __name__ == "__main__":
    get_num()
    main()
